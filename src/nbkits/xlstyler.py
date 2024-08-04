from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Border, Alignment, Side, Font, PatternFill
from openpyxl.styles.alignment import horizontal_alignments, vertical_aligments
from openpyxl.utils import get_column_letter, column_index_from_string

from .utils import to_hex_color

_LINE_STYLES = [
    "thin",
    "medium",
    "thick",
    "double",
    "hair",
    "dashed",
    "dotted",
    "dashDot",
    "dashDotDot",
    "mediumDashDot",
    "mediumDashDotDot",
    "mediumDashed",
    "slantDashDot",
]


class ExcelSheetStyler:
    def __init__(self, sheet: Worksheet):
        self._sheet = sheet

    def column_width(
        self,
        width: float | int | list[float | int],
        col: list[int] | list[str] | None = None,
    ):
        sheet = self._sheet
        if col is None and isinstance(width, list):
            col = range(1, 1 + len(width))
        else:
            col = _parse_arg_cols(col=col, max_column=sheet.max_column)

        if isinstance(width, (list, tuple)):
            if len(width) != len(col):
                msg = (
                    "The lists, 'width' and 'cols' , must be of equal length:"
                    f" {len(width)} != {len(col)}"
                )
                raise ValueError(msg)

        elif isinstance(width, (int, float)):
            width = [width] * len(col)

        for i, w in zip(col, width):
            dim = sheet.column_dimensions[get_column_letter(i)]
            dim.width = w

        return self

    def row_height(
        self,
        height: float | int | list[float | int],
        row: list[int] | list[str] = None,
        skip_rows: list[int] = None,
        skip_header: int | None = None,
        skip_footer: int = None,
    ):
        sheet = self._sheet
        row = _parse_arg_rows(
            row=row,
            max_row=sheet.max_row,
            skip_rows=skip_rows,
            skip_header=skip_header,
            skip_footer=skip_footer,
        )

        if isinstance(height, (list, tuple)):
            if len(height) != len(row):
                msg = (
                    "The lists, 'height' and 'rows' , must be of equal length:"
                    f" {len(height)} != {len(row)}"
                )
                raise ValueError(msg)

        elif isinstance(height, (int, float)):
            height = [height] * len(row)

        for i, h in zip(row, height):
            self._sheet.row_dimensions[i].height = h

        return self

    def border(
        self,
        col=None,
        row=None,
        sides=None,
        t=None,
        l=None,  # noqa: E741
        b=None,
        r=None,
        h=None,
        v=None,
        u: bool | None = None,
        d: bool | None = None,
        ls="thin",
        c: str | None = None,
        skip_rows=None,
        skip_header: int | None = None,
        skip_footer=None,
    ):
        sheet = self._sheet
        max_column, max_row = sheet.max_column, sheet.max_row
        col = _parse_arg_cols(col=col, max_column=max_column)
        row = _parse_arg_rows(
            row=row,
            max_row=max_row,
            skip_rows=skip_rows,
            skip_header=skip_header,
            skip_footer=skip_footer,
        )

        side_opts = {}
        if sides is not None:
            if not isinstance(sides, str):
                raise ValueError(
                    "Argument 'sides' has invalid side flags. "
                    "The flag must be one of 't', 'r', 'b', "
                    "'l', 'h', 'v', 'u', and 'd'."
                )

            sides = sides.lower()
            if sides == "all":
                sides = "trblhv"
            elif sides == "outside":
                sides = "trbl"
            elif sides == "inside":
                sides = "hv"

            _unknowns = []
            for flag in sides:
                if flag in _side_flag_map:
                    side_opts[_side_flag_map[flag]] = True
                else:
                    _unknowns.append(flag)
            if _unknowns:
                _flags = ", ".join([f"'{c}'" for c in _unknowns])
                raise ValueError(
                    f"unknown side flag: {_flags}"
                    "The flag must be one of 't', 'r', 'b', "
                    "'l', 'h', 'v', 'u', and 'd'."
                )

        for value, attr_name in [
            (t, "top"),
            (r, "right"),
            (b, "bottom"),
            (l, "left"),
            (h, "horizontal"),
            (v, "vertical"),
        ]:
            _parse_border_side_option(side_opts, value, attr_name)

        if u is not None:
            if isinstance(u, bool):
                side_opts["diagonalUp"] = u
            else:
                raise ValueError(
                    "The diagnoal-up side argument 'u' must be a True or False"
                )

        if d is not None:
            if isinstance(d, bool):
                side_opts["diagonalDown"] = d
            else:
                raise ValueError(
                    "The diagnoal-down side argument 'd' must be a True or False"
                )

        _normalize_side_opts(side_opts, ls=ls, c=c)

        min_row, max_row = min(row), max(row)
        min_col, max_col = min(col), max(col)
        for i in row:
            for j in col:
                cell = sheet.cell(i, j)
                kwargs = _get_border_args(cell.border)

                middle_sides = set(_OUTSIDES)
                if i == min_row:
                    _update_kwargs(kwargs, side_opts, "top")
                    middle_sides.remove("top")

                if i == max_row:
                    _update_kwargs(kwargs, side_opts, "bottom")
                    middle_sides.remove("bottom")

                if j == min_col:
                    _update_kwargs(kwargs, side_opts, "left")
                    middle_sides.remove("left")

                if j == max_col:
                    _update_kwargs(kwargs, side_opts, "right")
                    middle_sides.remove("right")

                for side in middle_sides:
                    if side in ["top", "bottom"]:
                        if "horizontal" in side_opts:
                            kwargs[side] = side_opts["horizontal"]
                    elif side in ["right", "left"]:
                        if "vertical" in side_opts:
                            kwargs[side] = side_opts["vertical"]

                cell.border = Border(**kwargs)

        return self

    def format(
        self,
        col=None,
        row=None,
        ha=None,
        va=None,
        wrap_text=None,
        indent=0,
        text_rotation=None,
        shrink_to_fit=None,
        number_format=None,
        family: str | None = None,
        size: float | int | None = None,
        color: str | None = None,
        background_color: str | None = None,
        bold: bool | None = None,
        italic: bool | None = None,
        strike: bool = None,
        baseline: bool | None = None,
        superscript: bool | None = None,
        subscript: bool | None = None,
        underline: bool | str | None = None,
        skip_rows=None,
        skip_header: int | None = None,
        skip_footer=None,
    ):
        sheet = self._sheet
        max_column, max_row = sheet.max_column, sheet.max_row
        col = _parse_arg_cols(col=col, max_column=max_column)
        row = _parse_arg_rows(
            row=row,
            max_row=max_row,
            skip_rows=skip_rows,
            skip_header=skip_header,
            skip_footer=skip_footer,
        )

        if ha and ha not in horizontal_alignments:
            values = ", ".join([f"'{v}'" for v in horizontal_alignments])
            raise ValueError(f"should be one of these values: {values}")

        if va and va not in vertical_aligments:
            values = ", ".join([f"'{v}'" for v in vertical_aligments])
            raise ValueError(f"should be one of these values: {values}")

        font_kwargs = {}
        if family is not None and isinstance(family, str):
            font_kwargs["name"] = family

        if size is not None and isinstance(size, (float, int)):
            font_kwargs["size"] = size

        if color is not None and isinstance(color, str):
            font_kwargs["color"] = to_hex_color(color)

        if bold is not None and isinstance(bold, bool):
            font_kwargs["bold"] = bold

        if italic is not None and isinstance(italic, bool):
            font_kwargs["italic"] = italic

        if strike is not None and isinstance(strike, bool):
            font_kwargs["strike"] = strike

        if baseline is not None and isinstance(baseline, bool):
            font_kwargs["vertAlign"] = "baseline"

        if superscript is not None and isinstance(superscript, bool):
            font_kwargs["vertAlign"] = "superscript"

        if subscript is not None and isinstance(subscript, bool):
            font_kwargs["vertAlign"] = "subscript"

        if underline is not None:
            if isinstance(underline, bool):
                underline = "single"
            elif isinstance(underline, str):
                if underline not in _UNDERLINE_OPTIONS:
                    _opts = ", ".join(f"'{c}'" for c in _UNDERLINE_OPTIONS)
                    raise ValueError(
                        "argument 'underline' must be a bool True/False or "
                        f"one of these values: {_opts}"
                    )
            font_kwargs["underline"] = underline

        font = Font(**font_kwargs) if font_kwargs else None

        fill = None
        if background_color is not None and isinstance(background_color, str):
            background_color = to_hex_color(background_color)
            fill = PatternFill(fill_type="solid", start_color=background_color)

        for i in row:
            for j in col:
                cell = sheet.cell(i, j)
                if ha or va or wrap_text or indent or shrink_to_fit or text_rotation:
                    cell.alignment = Alignment(
                        horizontal=ha,
                        vertical=va,
                        wrap_text=wrap_text,
                        indent=indent,
                        shrink_to_fit=shrink_to_fit,
                        text_rotation=text_rotation,
                    )

                if number_format is not None:
                    cell.number_format = number_format

                if font is not None:
                    cell.font = font

                if fill is not None:
                    cell.fill = fill

        return self

    def patten_fill(
        self,
        col: str | int | list[str] | list[int] | None = None,
        row: str | int | list[str] | list[int] | None = None,
        type: str | None = None,
        color: str | None = None,
        background_color: str | None = None,
        skip_rows=None,
        skip_header: int | None = None,
        skip_footer=None,
    ):
        sheet = self._sheet
        col = _parse_arg_cols(col=col, max_column=sheet.max_column)
        row = _parse_arg_rows(
            row=row,
            max_row=sheet.max_row,
            skip_rows=skip_rows,
            skip_header=skip_header,
            skip_footer=skip_footer,
        )

        kwargs = {}

        fill = None
        if type is None:
            kwargs["fill_type"] = "none"
        elif isinstance(type, str):
            kwargs["fill_type"] = type
        else:
            raise ValueError("argument `type` must be a str or None")

        if color is not None and isinstance(color, str):
            kwargs["start_color"] = to_hex_color(color)

        if background_color is not None and isinstance(background_color, str):
            kwargs["end_color"] = to_hex_color(background_color)

        fill = PatternFill(**kwargs)

        for i in row:
            for j in col:
                cell = sheet.cell(i, j)
                cell.fill = fill


_UNDERLINE_OPTIONS = ["single", "double", "singleAccounting", "doubleAccounting"]


def _update_kwargs(old_opts, side_opts, attr_name):
    if attr_name in side_opts:
        old_opts[attr_name] = side_opts[attr_name]


def _parse_border_side_option(side_opts, value, attr_name):
    if value is None:
        return

    if isinstance(value, str) and value in _LINE_STYLES:
        side_opts[attr_name] = value
        return

    if isinstance(value, bool):
        side_opts[attr_name] = value
        return

    opts = ", ".join([f"'{s}'" for s in _LINE_STYLES])
    raise ValueError(
        f"Invalid argument: '{attr_name}' must be either "
        f"True or False, or one of {opts}."
    )


_OUTSIDES = ["top", "right", "bottom", "left"]
_INSIDES = ["horizontal", "vertical"]

_side_flag_map = {
    "t": "top",
    "r": "right",
    "b": "bottom",
    "l": "left",
    "h": "horizontal",
    "v": "vertical",
    "u": "diagonalUp",
    "d": "diagonalDown",
}


def _normalize_side_opts(side_opts, ls=None, c=None):
    if isinstance(c, str):
        c = to_hex_color(c)

    if ls is None and c is None:
        side = Side(style="thin")
    elif ls is not None:
        if not (isinstance(ls, str) and ls in _LINE_STYLES):
            _flags = ", ".join([f"'{s}'" for s in _LINE_STYLES])
            raise ValueError(f"Invalid argument: 'ls' must be one of {_flags}.")

        side = Side(style=ls, color=c)
    else:
        side = Side(style="thin", color=c)  # 指定边线颜色但未指定边线样式，默认thin

    sides = set(["top", "right", "bottom", "left", "horizontal", "vertical"])
    for attr_name, value in side_opts.items():
        if attr_name in sides:
            if isinstance(value, bool):
                side_opts[attr_name] = side
            elif isinstance(value, str):
                side_opts[attr_name] = Side(style=value, color=c)

    if any(side_opts.get(s, False) for s in ["diagonalUp", "diagonalDown"]):
        side_opts["diagonal"] = side


def _get_border_args(border):
    """如果没有设置值则按默认值使用，如果设置了保留原值."""
    args = {}
    for side_name in _OUTSIDES:
        if hasattr(border, side_name):
            diagonal = getattr(border, side_name)
            if diagonal and diagonal.style is not None:
                args[side_name] = diagonal

    if hasattr(border, "diagonal"):
        diagonal = getattr(border, "diagonal")
        if diagonal and diagonal.style is not None:
            args["diagonal"] = diagonal

    for direction in ["diagonalUp", "diagonalDown"]:
        if hasattr(border, direction):
            args[direction] = getattr(border, direction)

    return args


def _parse_column_range(s: str):
    segs = s.split(":")
    n_segs = len(segs)
    if n_segs == 1:
        try:
            col_idx = column_index_from_string(s)
            return [col_idx]
        except ValueError:
            pass
        raise ValueError(s)
    elif n_segs == 2:
        start, stop = segs
        try:
            start = column_index_from_string(start)
            stop = column_index_from_string(stop)
            return list(range(start, stop + 1))
        except ValueError:
            pass
        raise ValueError(s)
    else:
        raise ValueError(s)


def _parse_index_list(idx):
    if isinstance(idx, int):
        return [idx]
    elif isinstance(idx, range):
        return list(idx)

    _idxs = []
    _invalids = []

    for c in idx:
        if isinstance(c, int):
            _idxs.append(c)
        elif isinstance(c, range):
            _idxs += list(c)
        else:
            _invalids.append(c)
    if _invalids:
        msg = "， ".join([f"'{c}'" for c in _invalids])
        raise ValueError(msg)

    return _keep_unqiue_sorted(_idxs)


def _keep_unqiue_sorted(idxs):
    uniques, prev = [], None
    for i in sorted(idxs):
        if prev != i:
            prev = i
            uniques.append(i)
        else:
            continue
    return uniques


def _parse_arg_rows(
    row=None, max_row=None, skip_rows=None, skip_header=None, skip_footer=None
):
    if isinstance(row, (list, tuple)):
        row = _parse_index_list(row)
    elif isinstance(row, range):
        row = list(row)
    elif isinstance(row, int):
        row = [row]
    elif row is None:
        row = list(range(1, max_row + 1))
    else:
        raise ValueError(f"unknown rows='{repr(row)}'")

    if len(row) == 0:
        raise ValueError("no rows specified")

    if skip_rows is None:
        skip_rows = []

    if skip_header is not None and isinstance(skip_header, int):
        skip_rows += list(range(1, max_row + 1))[:skip_header]

    if skip_footer is not None and isinstance(skip_footer, int):
        skip_rows += list(range(1, max_row + 1))[(-skip_header):]

    if skip_rows:
        row = [r for r in row if r not in skip_rows]

    return row


def _parse_arg_cols(col=None, max_column=None):
    idxs = None
    if col is None:
        idxs = list(range(1, max_column + 1))

    elif isinstance(col, (list, tuple)):
        if all(isinstance(c, str) for c in col):
            idxs = []
            _invalids = []
            for c in col:
                try:
                    idxs += _parse_column_range(c)
                except ValueError as ex:
                    _invalids.append(c)
            if _invalids:
                msg = "， ".join([f"'{c}'" for c in _invalids])
                msg = f"invalid column names: {msg}"
                raise ValueError(msg)

            idxs = _keep_unqiue_sorted(idxs)
        else:
            idxs = _parse_index_list(col)

    elif isinstance(col, range):
        idxs = list(col)
    elif isinstance(col, int):
        idxs = [col]
    else:
        raise ValueError(f"unknown argument: cols={repr(col)}")

    if not idxs:
        raise ValueError("empty column list")

    return idxs


def xlstyle(sheet):
    return ExcelSheetStyler(sheet)
