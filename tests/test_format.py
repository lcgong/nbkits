import openpyxl
from nbkits.xlstyler import xlstyle
from pathlib import Path

def main():
    wb = openpyxl.Workbook()
    ws = wb.active

    xlstyle(ws).column_width(20, col=range(2, 10))

    row = 2
    cols = ["b", "g", "r", "c", "m", "y", "k", "w"]
    xlstyle(ws).row_height(18, row=range(row, row + 2))
    for j, color in enumerate(cols):
        ws.cell(row, 2 + j, color)
        xlstyle(ws).format(
            row=row,
            col=2 + j,
            color=color,
            va="center",
            ha="center",
            background_color="k" if color == "w" else None,
        )
        ws.cell(row + 1, 2 + j, color)
        xlstyle(ws).format(
            row=row + 1,
            col=2 + j,
            background_color=color,
            color="w" if color == "k" else None,
            va="center",
            ha="center",
        )
        # xlstyle(ws).background(row=row+1, col=2 + j, color=color)

    row = 5
    types = ["none", "solid", "darkDown", "darkGray", "darkGrid", "darkHorizontal"]
    types += ["darkTrellis", "darkUp", "darkVertical", "gray0625", "gray125"]
    types += ["lightDown", "lightGray", "lightGrid", "lightHorizontal"]
    types += ["lightTrellis", "lightUp", "lightVertical", "mediumGray"]
    xlstyle(ws).row_height(18, row=range(row, row + 6))
    for j, typ in enumerate(types):
        if j == 8:
            row += 3

        j %= 8

        ws.cell(row + 1, 2 + j, typ)
        (
            xlstyle(ws)
            .format(row=row + 1, col=2 + j, va="center", ha="center", bold=True)
            .border(row=range(row, row + 3), col=2 + j, sides="outside", c="k")
            .patten_fill(
                row=range(row, row + 3),
                col=2 + j,
                type=typ,
                color="c",
                background_color="E0E0E0",
            )
        )

    row = 12
    xlstyle(ws).row_height(40, row=row)
    opts1 = ["top", "center", "bottom", "center", "center", "center"]
    opts2 = ["center", "center", "center", "left", "center", "right"]
    cols = [2, 3, 4, 6, 7, 8]
    for j, va, ha in zip(cols, opts1, opts2):
        ws.cell(row, j, f"{va}-{ha}")
        xlstyle(ws).format(row=row, col=j, va=va, ha=ha, bold=True)

    row = 12
    cols = range(2, 6)
    font_sizes = ["single", "double", "singleAccounting", "doubleAccounting"]
    xlstyle(ws).row_height(40, row=row)
    for j, sz in zip(cols, font_sizes):
        ws.cell(row, j, sz)
        xlstyle(ws).format(
            row=row, col=j, va="center", ha="center", underline=sz, bold=True
        )

    row = 14
    font_sizes = ["bold", "italic", "strike", "underline"]
    font_sizes += ["baseline", "superscript", "subscript"]
    cols = range(2, 2 + len(font_sizes))
    xlstyle(ws).row_height(40, row=row)
    for j, sz in zip(cols, font_sizes):
        ws.cell(row, j, sz)
        kwargs = {sz: True}
        xlstyle(ws).format(row=row, col=j, va="center", ha="center", **kwargs)

    
    row = 16
    font_sizes = [6, 8, 10, 12, 14, 16]
    xlstyle(ws).row_height(40, row=row)
    for j, sz in enumerate(font_sizes):
        ws.cell(row, 2 + j, f"楷体 {sz}pt")
        xlstyle(ws).format(row=row, col=2 + j, va="center", ha="center", family="楷体", size=sz)


    test_dir = Path(__file__).parent
    wb.save(test_dir / "test_format.xlsx")


if __name__ == "__main__":
    main()
